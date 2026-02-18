#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Týdenní export HubSpot dealů do Excelu (dynamický cutoff):
- Cutoff: closedate < (neděle minulého týdne + 18 měsíců).
- Agregace: součet 'amount' podle (deal owner × deal stage).
- Excel: každý owner = samostatný list; řádky = stage; sloupce = jednotlivé týdny (minulý týden).

Výstup:
  /Users/ladis/Library/CloudStorage/OneDrive-Sdílenéknihovny–Dateios.r.o/Dateio - TapiX/Sales/Leadgen team/Python Automatizations/Sales Report/HubSpot_Deals_By_Stage_Dynamic_2026.xlsx
"""

import os
import time
import requests
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple
from dotenv import load_dotenv
import openpyxl
from dateutil.relativedelta import relativedelta

try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except ImportError:
    from backports.zoneinfo import ZoneInfo

# ===== Konfigurace =====
BASE_DIR = Path(
    "/Users/ladis/Library/CloudStorage/OneDrive-Sdílenéknihovny–Dateios.r.o/Dateio - TapiX/Sales/Leadgen team/Python Automatizations/Sales Report"
)
EXCEL_PATH = BASE_DIR / "HubSpot_Deals_By_Stage_Dynamic_2026.xlsx"
LOCAL_TZ = "Europe/Prague"
DEBUG_MAX_PAGES = None  # např. 2 při ladění


# ===== Pomocné funkce =====
def load_token() -> str:
    project_dir = Path(__file__).parent
    load_dotenv(dotenv_path=project_dir / ".env")
    token = os.getenv("HUBSPOT_TOKEN")
    if not token:
        raise RuntimeError("Chybí HUBSPOT_TOKEN v .env (HUBSPOT_TOKEN=pat-...)")
    return token


def hs_headers(token: str) -> Dict[str, str]:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def backoff_sleep(attempt: int):
    time.sleep(min(2 ** attempt, 32))


# ===== HubSpot API =====
def get_all_owners(token: str) -> Dict[str, str]:
    url = "https://api.hubapi.com/crm/v3/owners/"
    owners_map: Dict[str, str] = {}
    params = {"limit": 100, "archived": "false"}
    attempt = 0
    while True:
        resp = requests.get(url, headers=hs_headers(token), params=params)
        if resp.status_code == 429 or 500 <= resp.status_code < 600:
            attempt += 1
            backoff_sleep(attempt)
            continue
        resp.raise_for_status()
        data = resp.json()
        for o in data.get("results", []):
            owner_id = str(o.get("id"))
            name = (
                f"{o.get('firstName', '')} {o.get('lastName', '')}".strip()
                or o.get("email", f"Owner {owner_id}")
            )
            owners_map[owner_id] = name
        next_after = data.get("paging", {}).get("next", {}).get("after")
        if next_after:
            params["after"] = next_after
        else:
            break
    return owners_map


def get_stage_label_map(token: str) -> Tuple[Dict[str, str], List[str]]:
    url = "https://api.hubapi.com/crm/v3/pipelines/deals"
    attempt = 0
    while True:
        resp = requests.get(url, headers=hs_headers(token))
        if resp.status_code == 429 or 500 <= resp.status_code < 600:
            attempt += 1
            backoff_sleep(attempt)
            continue
        resp.raise_for_status()
        data = resp.json()
        break
    stage_label_map: Dict[str, str] = {}
    default_order: List[str] = []
    for pipe in data.get("results", []):
        stages = pipe.get("stages", [])
        if not default_order:
            default_order = [s.get("label", s.get("id")) for s in stages]
        for s in stages:
            stage_label_map[s.get("id")] = s.get("label", s.get("id"))
    return stage_label_map, default_order


def fetch_deals(token: str, cutoff_epoch_ms: int) -> List[dict]:
    url = "https://api.hubapi.com/crm/v3/objects/deals/search"
    body = {
        "filterGroups": [
            {
                "filters": [
                    {
                        "propertyName": "closedate",
                        "operator": "LT",
                        "value": cutoff_epoch_ms,
                    }
                ]
            }
        ],
        "properties": ["dealstage", "amount", "hubspot_owner_id", "closedate", "pipeline"],
        "limit": 100,
        "sorts": [{"propertyName": "closedate", "direction": "DESCENDING"}],
    }
    all_deals: List[dict] = []
    pages = 0
    after = None
    attempt = 0
    while True:
        if after:
            body["after"] = after
        resp = requests.post(url, headers=hs_headers(token), json=body)
        if resp.status_code == 429 or 500 <= resp.status_code < 600:
            attempt += 1
            backoff_sleep(attempt)
            continue
        resp.raise_for_status()
        data = resp.json()
        attempt = 0
        results = data.get("results", [])
        all_deals.extend(results)
        pages += 1
        if DEBUG_MAX_PAGES and pages >= DEBUG_MAX_PAGES:
            break
        after = data.get("paging", {}).get("next", {}).get("after")
        if not after:
            break
    return all_deals


# ===== Agregace a Excel =====
def aggregate_amounts_by_owner_and_stage(
    deals: List[dict],
    owners_map: Dict[str, str],
    stage_label_map: Dict[str, str],
) -> Dict[str, Dict[str, float]]:
    data: Dict[str, Dict[str, float]] = {}
    for d in deals:
        props = d.get("properties", {}) or {}
        stage_id = props.get("dealstage")
        stage_label = stage_label_map.get(stage_id, "Unknown stage")
        amount = props.get("amount")
        try:
            val = float(amount) if amount not in (None, "") else 0.0
        except ValueError:
            val = 0.0
        owner_id = props.get("hubspot_owner_id")
        owner_name = owners_map.get(str(owner_id), "Unassigned")
        data.setdefault(owner_name, {}).setdefault(stage_label, 0.0)
        data[owner_name][stage_label] += val
    return data


def previous_week_label(now_local: datetime) -> Tuple[str, datetime, datetime]:
    monday_this_week = now_local - timedelta(days=now_local.weekday())
    monday_prev = monday_this_week - timedelta(days=7)
    sunday_prev = monday_prev + timedelta(days=6)
    iso = monday_prev.isocalendar()
    label = f"{iso.year}-W{iso.week:02d} ({monday_prev.date()}—{sunday_prev.date()})"
    return label, monday_prev, sunday_prev


def ensure_sheet(
    wb: openpyxl.Workbook, title: str
) -> openpyxl.worksheet.worksheet.Worksheet:
    safe = title
    for ch in r"\/?*[]:":
        safe = safe.replace(ch, " ")
    safe = safe.strip() or "Unassigned"
    if safe in wb.sheetnames:
        return wb[safe]
    return wb.create_sheet(title=safe[:31])


def write_snapshot_to_excel(
    excel_path: Path,
    week_label: str,
    data_by_owner: Dict[str, Dict[str, float]],
    default_stage_order: List[str],
):
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        wb = openpyxl.load_workbook(excel_path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            del wb["Sheet"]

    for owner_name, stage_sums in data_by_owner.items():
        ws = ensure_sheet(wb, owner_name)
        if ws.max_row < 1 or ws["A1"].value != "Stage":
            ws["A1"] = "Stage"

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if week_label in headers:
            week_col = headers.index(week_label) + 1
        else:
            week_col = ws.max_column + 1
            ws.cell(row=1, column=week_col, value=week_label)

        existing_rows: Dict[str, int] = {}
        for r in range(2, ws.max_row + 1):
            lbl = ws.cell(row=r, column=1).value
            if lbl:
                existing_rows[lbl] = r

        next_row = ws.max_row + 1
        for stage in default_stage_order:
            if stage not in existing_rows:
                ws.cell(row=next_row, column=1, value=stage)
                existing_rows[stage] = next_row
                next_row += 1

        for stage in stage_sums.keys():
            if stage not in existing_rows:
                ws.cell(row=next_row, column=1, value=stage)
                existing_rows[stage] = next_row
                next_row += 1

        for stage, amount_sum in stage_sums.items():
            r = existing_rows[stage]
            ws.cell(row=r, column=week_col, value=amount_sum)

    wb.save(excel_path)


# ===== Hlavní běh =====
def main():
    token = load_token()
    now_local = datetime.now(ZoneInfo(LOCAL_TZ))
    week_label, monday_prev, sunday_prev = previous_week_label(now_local)

    # Cutoff = neděle minulého týdne + 18 měsíců (lokální čas → epoch ms)
    cutoff_local = (sunday_prev.replace(tzinfo=ZoneInfo(LOCAL_TZ))) + relativedelta(months=+18)
    cutoff_ms = int(cutoff_local.timestamp() * 1000)

    owners_map = get_all_owners(token)
    stage_label_map, default_stage_order = get_stage_label_map(token)
    deals = fetch_deals(token, cutoff_ms)
    data_by_owner = aggregate_amounts_by_owner_and_stage(deals, owners_map, stage_label_map)

    print("Ukládám do:", EXCEL_PATH)
    write_snapshot_to_excel(EXCEL_PATH, week_label, data_by_owner, default_stage_order)

    print(f"Hotovo. Zapsán snapshot pro {week_label}")
    print(f"Cutoff pro tento týden byl {cutoff_local.date()}")
    print(f"Soubor: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
