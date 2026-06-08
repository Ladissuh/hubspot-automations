import os
import time
import datetime as dt
from typing import Dict, List, Optional, Tuple, Iterable

import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import (
    CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties,
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from dotenv import load_dotenv

# =========================
# CONFIG
# =========================
from pathlib import Path

PROJECT_DIR = Path(__file__).parent
ENV_PATH = PROJECT_DIR / ".env"      # lokálně může existovat, v GitHubu nevadí
OUTPUT_DIR = Path("outputs")         # GitHub-friendly

BASE_URL = "https://api.hubapi.com"

DEFAULT_PRODUCTS = ["Tapix", "EcoTrack", "ATM Nearby", "Labelling", "OpenData", "Subscription"]

# ✅ přidali jsme owner_* sloupce
SHEET_HEADERS = [
    "snapshot_week_start",   # pondělí daného týdne (idempotentní snapshot)
    "deal_id",
    "deal_name",
    "company_id",
    "company_name",
    "product_option",
    "product_raw",
    "pipeline_id",
    "pipeline_label",
    "dealstage_id",
    "dealstage_label",
    "amount",
    "closedate",
    "createdate",
    "hs_lastmodifieddate",
    "owner_id",
    "owner_name",
    "owner_email",
    "deal_url",
]

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=14)
WRAP = Alignment(vertical="center", wrap_text=True)

# =========================
# WEIGHTED PIPELINE CONFIG
# =========================
# weighted_amount = amount * win-probability of the deal's stage.
# "Tapix" values are the real HubSpot stage probabilities. "Leads" and
# "Account management" are reasonable estimates — edit them here and re-run
# (or re-run build_dashboard on an existing file) to refresh the dashboard.
STAGE_PROBABILITY = {
    "Tapix": {
        "Qualify": 0.09, "Discover": 0.12, "Validate": 0.20, "Decide": 0.45,
        "Commit": 0.90, "Tech intro": 0.95, "Implementation": 0.95, "Testing": 0.95,
        "Won": 1.00, "Lost": 0.00,
    },
    "Leads": {
        "To be contacted": 0.05, "Lead Engaged": 0.10, "Intro meeting agreed": 0.20,
        "Meeting negotiation / contacted": 0.30, "Qualified -> Deal": 0.50,
        "Awaiting confirmation / Not Now": 0.15, "Not a lead": 0.00,
    },
    "Account management (cross-sell / upsell)": {
        "Grow interest": 0.15, "Introduce/Pitch": 0.25, "Wait": 0.10,
        "Upsell Qualified -> Deal": 0.50, "Closed won": 1.00, "Closed lost": 0.00,
    },
}
# Funnel display order per pipeline (stages found in data but not listed here
# are appended automatically).
STAGE_ORDER = {
    "Tapix": ["Qualify", "Discover", "Validate", "Decide", "Commit",
              "Tech intro", "Implementation", "Testing", "Won", "Lost"],
    "Leads": ["To be contacted", "Lead Engaged", "Intro meeting agreed",
              "Meeting negotiation / contacted", "Qualified -> Deal",
              "Awaiting confirmation / Not Now", "Not a lead"],
    "Account management (cross-sell / upsell)": [
        "Grow interest", "Introduce/Pitch", "Wait",
        "Upsell Qualified -> Deal", "Closed won", "Closed lost"],
}
DEFAULT_PROBABILITY = 0.50            # any stage not configured above
ALL_PRODUCTS_LABEL = "All products"   # synthetic, deal-deduplicated aggregate


def _norm(s) -> str:
    return str(s if s is not None else "").strip()


def stage_probability(pipeline_label, stage_label) -> float:
    """Win probability for a (pipeline, stage); falls back to DEFAULT_PROBABILITY."""
    table = STAGE_PROBABILITY.get(_norm(pipeline_label), {})
    s = _norm(stage_label)
    if s in table:
        return table[s]
    low = {k.lower(): v for k, v in table.items()}
    return low.get(s.lower(), DEFAULT_PROBABILITY)


def stage_class(prob: float) -> str:
    """open / won / lost classification derived from probability."""
    if prob >= 1.0:
        return "won"
    if prob <= 0.0:
        return "lost"
    return "open"


# =========================
# REVENUE-ATTRIBUTION SPLIT
# =========================
# When a deal touches multiple of the 5 attribution products below, its amount
# is split among them according to the matching N-tuple ratio. Labelling (and
# any other product outside this scheme) always shows the FULL deal amount on
# its own sheet — it is never split. The split applies to per-product chart
# and KPI views only; the "All products" dropdown in the dashboard keeps the
# deduplicated whole-deal amount.

ATTR_ABBREV = {
    "Tapix":        "T",
    "Subscription": "R",
    "EcoTrack":     "E",
    "ATM Nearby":   "A",
    "OpenData":     "O",
}

ATTR_SPLITS = {
    # ---- 2-tuples ----
    frozenset({"T", "R"}):           {"T": 0.667, "R": 0.333},
    frozenset({"T", "E"}):           {"T": 0.588, "E": 0.412},
    frozenset({"T", "A"}):           {"T": 0.845, "A": 0.155},
    frozenset({"T", "O"}):           {"T": 0.694, "O": 0.306},
    frozenset({"R", "E"}):           {"R": 0.417, "E": 0.583},
    frozenset({"R", "A"}):           {"R": 0.731, "A": 0.269},
    frozenset({"R", "O"}):           {"R": 0.532, "O": 0.468},
    frozenset({"E", "A"}):           {"E": 0.792, "A": 0.208},
    frozenset({"E", "O"}):           {"E": 0.614, "O": 0.386},
    frozenset({"A", "O"}):           {"A": 0.295, "O": 0.705},
    # ---- 3-tuples ----
    frozenset({"T", "R", "E"}):      {"T": 0.455, "R": 0.227, "E": 0.318},
    frozenset({"T", "R", "A"}):      {"T": 0.594, "R": 0.297, "A": 0.109},
    frozenset({"T", "R", "O"}):      {"T": 0.515, "R": 0.258, "O": 0.227},
    frozenset({"T", "E", "A"}):      {"T": 0.531, "E": 0.372, "A": 0.098},
    frozenset({"T", "E", "O"}):      {"T": 0.467, "E": 0.327, "O": 0.206},
    frozenset({"T", "A", "O"}):      {"T": 0.616, "A": 0.113, "O": 0.271},
    frozenset({"R", "E", "A"}):      {"R": 0.361, "E": 0.506, "A": 0.133},
    frozenset({"R", "E", "O"}):      {"R": 0.305, "E": 0.427, "O": 0.268},
    frozenset({"R", "A", "O"}):      {"R": 0.445, "A": 0.163, "O": 0.392},
    frozenset({"E", "A", "O"}):      {"E": 0.529, "A": 0.139, "O": 0.332},
    # ---- 4-tuples ----
    frozenset({"T", "R", "E", "A"}): {"T": 0.420, "R": 0.210, "E": 0.294, "A": 0.077},
    frozenset({"T", "R", "E", "O"}): {"T": 0.379, "R": 0.189, "E": 0.265, "O": 0.167},
    frozenset({"T", "R", "A", "O"}): {"T": 0.471, "R": 0.235, "A": 0.086, "O": 0.207},
    frozenset({"T", "E", "A", "O"}): {"T": 0.430, "E": 0.301, "A": 0.079, "O": 0.189},
    frozenset({"R", "E", "A", "O"}): {"R": 0.274, "E": 0.384, "A": 0.101, "O": 0.241},
    # ---- 5-tuple ----
    frozenset({"T", "R", "E", "A", "O"}):
        {"T": 0.354, "R": 0.177, "E": 0.248, "A": 0.065, "O": 0.156},
}


def attribution_split_factor(product_set, target_product) -> float:
    """Share of a deal's amount assigned to ``target_product`` based on its
    full product set. Labelling (or any product outside ATTR_ABBREV) returns
    1.0 — it is never split. Among the attribution-5 products, the split is
    based on the deal's intersection with ATTR_ABBREV (Labelling is filtered
    out before computing the ratio)."""
    if target_product not in ATTR_ABBREV:
        return 1.0
    abbr_target = ATTR_ABBREV[target_product]
    abbr_set = frozenset(ATTR_ABBREV[p] for p in product_set if p in ATTR_ABBREV)
    if len(abbr_set) <= 1:
        return 1.0
    ratios = ATTR_SPLITS.get(abbr_set)
    if not ratios:                              # safety: unknown combo
        return 1.0 / len(abbr_set)
    return ratios.get(abbr_target, 0.0)


# =========================
# HELPERS
# =========================

def hubspot_request(token: str, method: str, path: str, *, params=None, json=None, retries: int = 5):
    url = f"{BASE_URL}{path}"
    headers = {"authorization": f"Bearer {token}", "Content-Type": "application/json"}

    last = None
    for attempt in range(retries):
        r = requests.request(method, url, headers=headers, params=params, json=json, timeout=60)
        last = r

        if r.status_code in (429, 500, 502, 503, 504):
            time.sleep(min(2 ** attempt, 20))
            continue

        r.raise_for_status()
        # některé endpointy vrací [] (list), jiné dict
        if r.text and r.text.strip():
            return r.json()
        return {}

    if last is not None:
        last.raise_for_status()
    raise RuntimeError("hubspot_request failed without response")


def week_start_iso(d: Optional[dt.date] = None) -> str:
    """Vrátí ISO datum pondělí aktuálního týdne (YYYY-MM-DD)."""
    if d is None:
        d = dt.date.today()
    monday = d - dt.timedelta(days=d.weekday())
    return monday.isoformat()


def excel_safe_sheet_name(name: str) -> str:
    return name[:31]


def chunked(items: List[str], size: int) -> Iterable[List[str]]:
    for i in range(0, len(items), size):
        yield items[i:i + size]


def ensure_sheet_headers(ws):
    for c, h in enumerate(SHEET_HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def create_new_workbook(path: str, products: List[str]) -> Workbook:
    """Vytvoří nový workbook (uživatel smaže starý)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Weekly HubSpot Deals by Product — Summary"
    ws["A1"].font = TITLE_FONT
    ws.sheet_view.showGridLines = False

    for p in products:
        sh = wb.create_sheet(title=excel_safe_sheet_name(p))
        sh.append(SHEET_HEADERS)
        ensure_sheet_headers(sh)

    wb.save(path)
    return wb


def find_product_property_name(token: str, label: str, explicit_name: Optional[str]) -> str:
    if explicit_name:
        return explicit_name

    props = hubspot_request(token, "GET", "/crm/v3/properties/deals")
    for p in props.get("results", []):
        if (p.get("label") or "").strip().lower() == label.strip().lower():
            return p["name"]

    raise RuntimeError(
        f"Nenašel jsem deal property s labelem '{label}'. "
        f"Dej do .env PRODUCT_PROPERTY_NAME=<internal_name>."
    )


def get_product_options_map(token: str, property_name: str) -> Dict[str, str]:
    prop = hubspot_request(token, "GET", f"/crm/v3/properties/deals/{property_name}")
    options = prop.get("options", []) or []
    return {o.get("value"): o.get("label", o.get("value")) for o in options if o.get("value")}


def get_pipelines_map(token: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    data = hubspot_request(token, "GET", "/crm/v3/pipelines/deals")
    pipeline_label = {}
    stage_label = {}

    for p in data.get("results", []):
        pid = p.get("id")
        pipeline_label[pid] = p.get("label", pid)
        for s in p.get("stages", []) or []:
            sid = s.get("id")
            stage_label[sid] = s.get("label", sid)

    return pipeline_label, stage_label


def list_all_deals(token: str, properties: List[str]) -> List[dict]:
    deals = []
    after = None

    while True:
        params = {
            "limit": 100,
            "properties": ",".join(properties),
            "archived": "false",
        }
        if after is not None:
            params["after"] = after

        page = hubspot_request(token, "GET", "/crm/v3/objects/deals", params=params)
        deals.extend(page.get("results", []))

        nxt = (page.get("paging") or {}).get("next") or {}
        after = nxt.get("after")
        if not after:
            break

    return deals


def split_multicheckbox(value: Optional[str]) -> List[str]:
    if not value:
        return []
    return [p.strip() for p in value.split(";") if p.strip()]


def get_owners_map(token: str) -> Dict[str, Dict[str, str]]:
    """
    owner_id -> {"name": "...", "email": "..."}
    Zkouší v3 endpoint; fallback legacy v2.
    """
    candidate_paths = ["/crm/v3/owners", "/owners/v2/owners"]
    data = None
    last_err = None
    for path in candidate_paths:
        try:
            data = hubspot_request(token, "GET", path)
            break
        except Exception as e:
            last_err = e
            data = None

    if data is None:
        raise RuntimeError(f"Nepodařilo se stáhnout owners. Poslední chyba: {last_err}")

    results = data.get("results", []) if isinstance(data, dict) else data

    owners: Dict[str, Dict[str, str]] = {}
    for o in results:
        oid = str(o.get("id") or "")
        if not oid:
            continue
        first = (o.get("firstName") or "").strip()
        last = (o.get("lastName") or "").strip()
        email = (o.get("email") or "").strip()
        name = (f"{first} {last}").strip() or email or oid
        owners[oid] = {"name": name, "email": email}

    return owners


def batch_read_deal_company_primary(token: str, deal_ids: List[str]) -> Dict[str, str]:
    """
    deal_id -> primary_company_id (nebo první associated)
    """
    if not deal_ids:
        return {}

    candidate_paths = [
        "/crm/v4/associations/deals/companies/batch/read",
        "/crm/v4/associations/deal/companies/batch/read",
        "/crm/v4/associations/deals/company/batch/read",
    ]

    result: Dict[str, str] = {}

    for batch in chunked(deal_ids, 1000):
        payload = {"inputs": [{"id": str(did)} for did in batch]}

        last_err = None
        data = None
        for path in candidate_paths:
            try:
                data = hubspot_request(token, "POST", path, json=payload)
                break
            except Exception as e:
                last_err = e
                data = None

        if data is None:
            raise RuntimeError(f"Nepodařilo se stáhnout associations deal->company. Poslední chyba: {last_err}")

        for rec in data.get("results", []):
            deal_id = str((rec.get("from") or {}).get("id"))
            tos = rec.get("to") or []

            primary_company_id = None
            fallback_first = None

            for t in tos:
                cid = str(t.get("toObjectId"))
                if not fallback_first:
                    fallback_first = cid

                assoc_types = t.get("associationTypes") or []
                if any((a.get("label") or "").lower() == "primary" for a in assoc_types):
                    primary_company_id = cid
                    break

            result[deal_id] = primary_company_id or fallback_first or ""

    return result


def batch_read_company_names(token: str, company_ids: List[str]) -> Dict[str, str]:
    """company_id -> company_name"""
    if not company_ids:
        return {}

    names: Dict[str, str] = {}
    for batch in chunked(company_ids, 100):
        payload = {
            "inputs": [{"id": str(cid)} for cid in batch],
            "properties": ["name"],
        }
        data = hubspot_request(
            token,
            "POST",
            "/crm/v3/objects/companies/batch/read",
            params={"archived": "false"},
            json=payload,
        )
        for r in data.get("results", []):
            cid = str(r.get("id"))
            props = r.get("properties") or {}
            names[cid] = props.get("name") or ""
    return names


def build_rows(
    deals: List[dict],
    snapshot_week_start: str,
    product_property_name: str,
    opt_map: Dict[str, str],
    pipeline_label: Dict[str, str],
    stage_label: Dict[str, str],
    products_interest: List[str],
    deal_to_company_id: Dict[str, str],
    company_id_to_name: Dict[str, str],
    owners_map: Dict[str, Dict[str, str]],
) -> Dict[str, List[List]]:
    want = {p.lower(): p for p in products_interest}
    rows_by_product: Dict[str, List[List]] = {p: [] for p in products_interest}

    # dedupe per product per snapshot: (product, deal_id)
    seen = set()

    for d in deals:
        deal_id = str(d.get("id"))
        props = d.get("properties") or {}

        dealname = props.get("dealname")
        amount = props.get("amount")
        closedate = props.get("closedate")
        createdate = props.get("createdate")
        lastmod = props.get("hs_lastmodifieddate")

        pipeline_id = props.get("pipeline")
        stage_id = props.get("dealstage")
        pipeline_lbl = pipeline_label.get(pipeline_id, pipeline_id)
        stage_lbl = stage_label.get(stage_id, stage_id)

        company_id = deal_to_company_id.get(deal_id, "") or ""
        company_name = company_id_to_name.get(company_id, "") if company_id else ""

        owner_id = str(props.get("hubspot_owner_id") or "")
        owner_name = owners_map.get(owner_id, {}).get("name", "") if owner_id else ""
        owner_email = owners_map.get(owner_id, {}).get("email", "") if owner_id else ""

        product_raw = props.get(product_property_name)
        product_values = split_multicheckbox(product_raw)
        product_labels = [opt_map.get(v, v) for v in product_values]

        for pl in product_labels:
            key = (pl or "").strip().lower()
            if key in want:
                sheet_product = want[key]
                if (sheet_product, deal_id) in seen:
                    continue
                seen.add((sheet_product, deal_id))

                deal_url = f"https://app.hubspot.com/contacts/deal/{deal_id}"

                rows_by_product[sheet_product].append([
                    snapshot_week_start,
                    deal_id,
                    dealname,
                    company_id,
                    company_name,
                    sheet_product,
                    product_raw,
                    pipeline_id,
                    pipeline_lbl,
                    stage_id,
                    stage_lbl,
                    amount,
                    closedate,
                    createdate,
                    lastmod,
                    owner_id,
                    owner_name,
                    owner_email,
                    deal_url
                ])

    return rows_by_product


# =========================
# ✅ ONLY CHANGE: WRITING LOGIC
# =========================

def snapshot_exists(ws, snapshot_week_start: str) -> bool:
    """
    Vrátí True, pokud už v sheetu existuje alespoň jeden řádek pro daný snapshot_week_start.
    """
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        if str(r[0]) == snapshot_week_start:
            return True
    return False


def replace_rows_for_snapshot(ws, rows: List[List], snapshot_week_start: str):
    """
    NOVÉ CHOVÁNÍ (doplňování po týdnech):
    - Pokud už snapshot pro tento týden existuje, NIC nepřepisuj (žádné mazání).
    - Pokud neexistuje, pouze přidej nové řádky (append).
    """
    if snapshot_exists(ws, snapshot_week_start):
        # týden už je zapsaný -> nech ho být (doplňování po týdnech)
        return

    for row in rows:
        ws.append(row)


# =========================
# SUMMARY (beze změn)
# =========================

def read_all_product_sheets(wb, products: List[str]) -> pd.DataFrame:
    frames = []
    for p in products:
        name = excel_safe_sheet_name(p)
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        data = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            r = list(r)
            if len(r) < len(SHEET_HEADERS):
                r += [None] * (len(SHEET_HEADERS) - len(r))
            data.append(r[:len(SHEET_HEADERS)])
        if data:
            frames.append(pd.DataFrame(data, columns=SHEET_HEADERS))

    if not frames:
        return pd.DataFrame(columns=SHEET_HEADERS)
    return pd.concat(frames, ignore_index=True)


def rewrite_summary_sheet(wb, products: List[str]):
    df = read_all_product_sheets(wb, products)
    ws = wb["Summary"]
    ws.delete_rows(1, ws.max_row)

    ws["A1"] = "Weekly HubSpot Deals by Product — Summary"
    ws["A1"].font = TITLE_FONT
    ws.sheet_view.showGridLines = False

    if df.empty:
        ws["A3"] = "No data yet. Run the script once to populate product sheets."
        return

    # typing / cleanup
    df["snapshot_week_start"] = df["snapshot_week_start"].astype(str)
    df["product_option"] = df["product_option"].astype(str)
    df["pipeline_label"] = df["pipeline_label"].astype(str)
    df["dealstage_label"] = df["dealstage_label"].astype(str)
    df["owner_name"] = df["owner_name"].fillna("").astype(str)
    df["amount_num"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)

    weeks = sorted(df["snapshot_week_start"].unique().tolist())
    prod_order = products[:]

    # -------------------------
    # Table 1: counts by product x week
    # -------------------------
    ws["A3"] = "Table 1: Deal count by product (weekly snapshots)"
    ws["A3"].font = Font(bold=True)

    start_row = 5
    ws.cell(row=start_row, column=1, value="Product").font = HEADER_FONT
    ws.cell(row=start_row, column=1).fill = HEADER_FILL
    ws.cell(row=start_row, column=1).alignment = WRAP

    for i, w in enumerate(weeks, start=2):
        c = ws.cell(row=start_row, column=i, value=w)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP
        ws.column_dimensions[get_column_letter(i)].width = 14

    counts = (
        df.groupby(["product_option", "snapshot_week_start"])["deal_id"]
          .nunique()
          .reset_index(name="deal_count")
    )
    count_pivot = counts.pivot(index="product_option", columns="snapshot_week_start", values="deal_count").fillna(0).astype(int)

    for r_i, p in enumerate(prod_order, start=start_row + 1):
        ws.cell(row=r_i, column=1, value=p)
        for c_i, w in enumerate(weeks, start=2):
            val = int(count_pivot.loc[p, w]) if (p in count_pivot.index and w in count_pivot.columns) else 0
            ws.cell(row=r_i, column=c_i, value=val)

    # -------------------------
    # Table 2: sum amount by product x week
    # -------------------------
    ws["A20"] = "Table 2: Total amount by product (weekly snapshots)"
    ws["A20"].font = Font(bold=True)

    start_row2 = 22
    ws.cell(row=start_row2, column=1, value="Product").font = HEADER_FONT
    ws.cell(row=start_row2, column=1).fill = HEADER_FILL
    ws.cell(row=start_row2, column=1).alignment = WRAP

    for i, w in enumerate(weeks, start=2):
        c = ws.cell(row=start_row2, column=i, value=w)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP

    sums = (
        df.groupby(["product_option", "snapshot_week_start"])["amount_num"]
          .sum()
          .reset_index(name="amount_sum")
    )
    sum_pivot = sums.pivot(index="product_option", columns="snapshot_week_start", values="amount_sum").fillna(0.0)

    for r_i, p in enumerate(prod_order, start=start_row2 + 1):
        ws.cell(row=r_i, column=1, value=p)
        for c_i, w in enumerate(weeks, start=2):
            val = float(sum_pivot.loc[p, w]) if (p in sum_pivot.index and w in sum_pivot.columns) else 0.0
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.number_format = "#,##0.00"

    # -------------------------
    # Table 3: tidy distribution incl. owner (Power BI friendly)
    # -------------------------
    ws["A37"] = "Table 3: Stage distribution incl. owner (tidy, good for Power BI)"
    ws["A37"].font = Font(bold=True)

    tidy_headers = [
        "snapshot_week_start",
        "product",
        "owner_name",
        "pipeline_label",
        "dealstage_label",
        "deal_count",
        "amount_sum",
    ]
    header_row = 39
    for c, h in enumerate(tidy_headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP

    tidy = (
        df.groupby(["snapshot_week_start", "product_option", "owner_name", "pipeline_label", "dealstage_label"])
          .agg(deal_count=("deal_id", "nunique"), amount_sum=("amount_num", "sum"))
          .reset_index()
          .rename(columns={"product_option": "product"})
          .sort_values(["snapshot_week_start", "product", "owner_name", "pipeline_label", "dealstage_label"])
    )

    row = header_row + 1
    for _, rec in tidy.iterrows():
        ws.cell(row=row, column=1, value=rec["snapshot_week_start"])
        ws.cell(row=row, column=2, value=rec["product"])
        ws.cell(row=row, column=3, value=rec["owner_name"])
        ws.cell(row=row, column=4, value=rec["pipeline_label"])
        ws.cell(row=row, column=5, value=rec["dealstage_label"])
        ws.cell(row=row, column=6, value=int(rec["deal_count"]))
        c7 = ws.cell(row=row, column=7, value=float(rec["amount_sum"]))
        c7.number_format = "#,##0.00"
        row += 1

    # basic formatting
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14


def _ordered_stages_for_pipeline(pipeline_label: str, data_stages) -> List[str]:
    """Ordered, data-exact stage labels for a pipeline.

    Uses STAGE_ORDER for funnel order; substitutes the exact string found in the
    data (so SUMIFS matches even when HubSpot labels carry trailing spaces) and
    appends any data stage not present in STAGE_ORDER.
    """
    cfg_order = STAGE_ORDER.get(_norm(pipeline_label), [])
    data_lookup = {}
    for s in data_stages:
        data_lookup.setdefault(_norm(s).lower(), s)
    ordered, used = [], set()
    for cfg_stage in cfg_order:
        k = _norm(cfg_stage).lower()
        ordered.append(data_lookup.get(k, cfg_stage))
        used.add(k)
    for s in sorted(data_stages):
        if _norm(s).lower() not in used:
            ordered.append(s)
            used.add(_norm(s).lower())
    return ordered


def build_dashboard(wb, products: List[str]):
    """Build/refresh the interactive 'Pipeline Dashboard' sheet and its hidden
    '_cfg' data sheet. Safe to re-run: both sheets are recreated from scratch."""
    df = read_all_product_sheets(wb, products)

    for nm in ("Pipeline Dashboard", "_cfg"):
        if nm in wb.sheetnames:
            del wb[nm]
    cfg = wb.create_sheet("_cfg")
    dash = wb.create_sheet("Pipeline Dashboard")
    cfg.sheet_state = "hidden"

    if df.empty:
        dash["A1"] = "Pipeline Dashboard"
        dash["A1"].font = TITLE_FONT
        dash["A3"] = "No data yet. Run the report to populate product sheets."
        return

    # ---- clean ----
    df = df.copy()
    df["snapshot_week_start"] = df["snapshot_week_start"].astype(str).str.strip()
    df["product_option"] = df["product_option"].astype(str).str.strip()
    df["pipeline_label"] = df["pipeline_label"].astype(str)
    df["dealstage_label"] = df["dealstage_label"].astype(str)
    df["owner_name"] = df["owner_name"].fillna("").astype(str)
    df["amount_num"] = pd.to_numeric(df["amount"], errors="coerce").fillna(0.0)

    # ---- revenue-attribution split for multi-product deals ----
    # per-product views use the SPLIT amount; "All products" (deduped) uses
    # the FULL amount so each deal counts once at its true value
    df["_deal_key"] = df["snapshot_week_start"].astype(str) + "|" + df["deal_id"].astype(str)
    deal_products_map = df.groupby("_deal_key")["product_option"].apply(frozenset).to_dict()
    df["amount_split"] = df.apply(
        lambda r: float(r["amount_num"]) * attribution_split_factor(
            deal_products_map.get(r["_deal_key"], frozenset()), r["product_option"]),
        axis=1,
    )

    weeks = sorted(df["snapshot_week_start"].unique().tolist())
    latest_week = weeks[-1]

    # ---- monthly buckets: each month represented by its LATEST weekly snapshot ----
    # (weekly snapshots are point-in-time states, so a month = its month-end snapshot;
    #  summing weeks within a month would double-count the same deals)
    month_to_weeks = {}
    for wk in weeks:
        month_to_weeks.setdefault(str(wk)[:7], []).append(wk)
    month_keys = sorted(month_to_weeks)
    month_rep = {mk: max(month_to_weeks[mk]) for mk in month_keys}

    def _month_label(mk):
        try:
            return dt.datetime.strptime(mk + "-01", "%Y-%m-%d").strftime("%B %Y")
        except Exception:
            return mk

    months = [_month_label(mk) for mk in month_keys]          # e.g. "March 2026"
    month_week = {_month_label(mk): month_rep[mk] for mk in month_keys}
    latest_month = months[-1]

    data_pipes = list(df["pipeline_label"].unique())
    pipe_pref = list(STAGE_ORDER.keys())
    pipelines = ([p for p in pipe_pref if p in data_pipes]
                 + sorted(p for p in data_pipes if p not in pipe_pref))

    rep_weeks = set(month_rep.values())          # the month-end snapshot weeks
    df_rep = df[df["snapshot_week_start"].isin(rep_weeks)]
    pipe_stages = {}
    pipe_data_stages = {}                        # stages charted: seen at a month-end
    for p in pipelines:
        ds_all = set(df.loc[df["pipeline_label"] == p, "dealstage_label"].unique())
        ds_rep = set(df_rep.loc[df_rep["pipeline_label"] == p, "dealstage_label"].unique())
        pipe_data_stages[p] = {_norm(s).lower() for s in ds_rep}
        pipe_stages[p] = _ordered_stages_for_pipeline(p, ds_all)
    M = max((len(v) for v in pipe_stages.values()), default=1)

    # ---- tidy source incl. deal-deduplicated "All products" ----
    # per_prod -> SPLIT amounts (per-product attribution); all_prod -> FULL amounts (deduped)
    g = ["snapshot_week_start", "product_option", "owner_name", "pipeline_label", "dealstage_label"]
    per_prod = (df.groupby(g)
                  .agg(deal_count=("deal_id", "nunique"), amount_sum=("amount_split", "sum"))
                  .reset_index()
                  .rename(columns={"product_option": "product"}))
    ded = df.drop_duplicates(subset=["snapshot_week_start", "deal_id"])
    all_prod = (ded.groupby(["snapshot_week_start", "owner_name", "pipeline_label", "dealstage_label"])
                   .agg(deal_count=("deal_id", "nunique"), amount_sum=("amount_num", "sum"))
                   .reset_index())
    all_prod["product"] = ALL_PRODUCTS_LABEL
    cols = ["snapshot_week_start", "product", "owner_name", "pipeline_label",
            "dealstage_label", "deal_count", "amount_sum"]
    src = pd.concat([all_prod[cols], per_prod[cols]], ignore_index=True)
    src["weighted_amount_sum"] = src.apply(
        lambda rec: float(rec["amount_sum"]) * stage_probability(rec["pipeline_label"], rec["dealstage_label"]),
        axis=1,
    )

    # ===== _cfg: DashData table (cols A:H) =====
    dd_headers = ["snapshot_week_start", "product", "owner_name", "pipeline_label",
                  "dealstage_label", "deal_count", "amount_sum", "weighted_amount_sum"]
    for c, h in enumerate(dd_headers, start=1):
        cfg.cell(row=1, column=c, value=h)
    r = 2
    for _, rec in src.iterrows():
        cfg.cell(row=r, column=1, value=str(rec["snapshot_week_start"]))
        cfg.cell(row=r, column=2, value=str(rec["product"]))
        cfg.cell(row=r, column=3, value=str(rec["owner_name"]))
        cfg.cell(row=r, column=4, value=str(rec["pipeline_label"]))
        cfg.cell(row=r, column=5, value=str(rec["dealstage_label"]))
        cfg.cell(row=r, column=6, value=int(rec["deal_count"]))
        cfg.cell(row=r, column=7, value=float(rec["amount_sum"]))
        cfg.cell(row=r, column=8, value=float(rec["weighted_amount_sum"]))
        r += 1
    n_rows = r - 1
    tbl = Table(displayName="DashData", ref=f"A1:H{n_rows}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    cfg.add_table(tbl)

    # ===== _cfg: StageCfg lookup (cols K:R) =====
    for c, h in zip(range(11, 19),
                    ["key", "pipeline_label", "dealstage_label", "stage_order",
                     "probability", "is_open", "is_won", "key2"]):
        cfg.cell(row=1, column=c, value=h)
    sr = 2
    for p in pipelines:
        open_idx = 0
        for idx, stage in enumerate(pipe_stages[p], start=1):
            prob = stage_probability(p, stage)
            cls = stage_class(prob)
            cfg.cell(row=sr, column=11, value=f"{p}|{idx}")
            cfg.cell(row=sr, column=12, value=p)
            cfg.cell(row=sr, column=13, value=stage)
            cfg.cell(row=sr, column=14, value=idx)
            cfg.cell(row=sr, column=15, value=prob)
            cfg.cell(row=sr, column=16, value=1 if cls == "open" else 0)
            cfg.cell(row=sr, column=17, value=1 if cls == "won" else 0)
            # key2 = chart column key; only OPEN stages that actually occur in the
            # data get one, so config-only phantom stages stay off the chart
            if cls == "open" and _norm(stage).lower() in pipe_data_stages[p]:
                open_idx += 1
                cfg.cell(row=sr, column=18, value=f"{p}|O{open_idx}")
            sr += 1
    S = sr - 1
    # the stacked chart is fixed to one pipeline (Tapix). Three pipelines with
    # different stage counts can't share one static chart without leaving empty
    # trailing series; Tapix is also the pipeline with confirmed probabilities.
    chart_pipeline = "Tapix" if "Tapix" in pipelines else (
        pipelines[0] if pipelines else "Tapix")
    M_OPEN = sum(
        1 for s in pipe_stages.get(chart_pipeline, [])
        if stage_class(stage_probability(chart_pipeline, s)) == "open"
        and _norm(s).lower() in pipe_data_stages.get(chart_pipeline, set())
    ) or 1

    # ===== _cfg: meta lists — months (S), products (T), pipelines (U),
    #       month -> representative week (V) — plus PipeWon (W, X) =====
    cfg.cell(row=1, column=19, value="months")
    cfg.cell(row=1, column=22, value="month_week")
    for i, mlabel in enumerate(months, start=2):
        cfg.cell(row=i, column=19, value=mlabel)
        cfg.cell(row=i, column=22, value=month_week[mlabel])
    cfg.cell(row=1, column=20, value="products")
    prod_list = [ALL_PRODUCTS_LABEL] + list(products)
    for i, pr in enumerate(prod_list, start=2):
        cfg.cell(row=i, column=20, value=pr)
    cfg.cell(row=1, column=21, value="pipelines")
    for i, p in enumerate(pipelines, start=2):
        cfg.cell(row=i, column=21, value=p)
    nM, NP = len(months), len(pipelines)

    cfg.cell(row=1, column=23, value="pipeline")
    cfg.cell(row=1, column=24, value="won_stage")
    for i, p in enumerate(pipelines, start=2):
        won = ""
        for stage in pipe_stages[p]:
            if stage_class(stage_probability(p, stage)) == "won":
                won = stage
                break
        cfg.cell(row=i, column=23, value=p)
        cfg.cell(row=i, column=24, value=won)

    # selected-pipeline won stage (Z1 / AA1)
    cfg.cell(row=1, column=26, value="won_stage_sel")
    cfg.cell(row=1, column=27,
             value=(f"=IFERROR(INDEX($X$2:$X${1+NP},"
                    f"MATCH('Pipeline Dashboard'!$C$5,$W$2:$W${1+NP},0)),\"\")"))
    # selected month -> representative week (Z2 / AA2)
    cfg.cell(row=2, column=26, value="sel_week")
    cfg.cell(row=2, column=27,
             value=(f"=IFERROR(INDEX($V$2:$V${1+nM},"
                    f"MATCH('Pipeline Dashboard'!$C$6,$S$2:$S${1+nM},0)),\"\")"))

    # ===== _cfg: result table (cols Z:AF, header row 3) =====
    res_hdr, res_first, res_last = 3, 4, 3 + M
    for c, h in zip(range(26, 33),
                    ["#", "Stage", "Raw amount", "Weighted amount", "Deals", "open", "won"]):
        cfg.cell(row=res_hdr, column=c, value=h)
    pref, prref, wref = ("'Pipeline Dashboard'!$C$5",
                         "'Pipeline Dashboard'!$C$4",
                         "_cfg!$AA$2")
    for k in range(1, M + 1):
        rr = res_hdr + k
        key = f"{pref}&\"|\"&$Z{rr}"
        cfg.cell(row=rr, column=26, value=k)
        cfg.cell(row=rr, column=27,
                 value=f"=IFERROR(INDEX($M$2:$M${1+S},MATCH({key},$K$2:$K${1+S},0)),\"\")")
        flt = (f"DashData[pipeline_label],{pref},"
               f"DashData[dealstage_label],$AA{rr},"
               f"DashData[product],{prref},"
               f"DashData[snapshot_week_start],{wref}")
        cfg.cell(row=rr, column=28,
                 value=f"=IF($AA{rr}=\"\",0,SUMIFS(DashData[amount_sum],{flt}))")
        cfg.cell(row=rr, column=29,
                 value=f"=IF($AA{rr}=\"\",0,SUMIFS(DashData[weighted_amount_sum],{flt}))")
        cfg.cell(row=rr, column=30,
                 value=f"=IF($AA{rr}=\"\",0,SUMIFS(DashData[deal_count],{flt}))")
        cfg.cell(row=rr, column=31,
                 value=f"=IFERROR(INDEX($P$2:$P${1+S},MATCH({key},$K$2:$K${1+S},0)),0)")
        cfg.cell(row=rr, column=32,
                 value=f"=IFERROR(INDEX($Q$2:$Q${1+S},MATCH({key},$K$2:$K${1+S},0)),0)")
        cfg.cell(row=rr, column=28).number_format = "#,##0"
        cfg.cell(row=rr, column=29).number_format = "#,##0"

    # ===== _cfg: stacked matrix — weighted amount, month (rows) x stage (cols) =====
    # each month uses its latest weekly snapshot; one column per OPEN stage
    stk_hdr, stk_first, stk_last = 3, 4, 3 + nM
    WK = 37  # col AK = month labels
    cfg.cell(row=stk_hdr, column=WK, value="Month")
    for i, mlabel in enumerate(months, start=1):
        cfg.cell(row=stk_hdr + i, column=WK, value=mlabel)
    cp_lit = '"' + chart_pipeline.replace('"', '""') + '"'   # chart fixed to Tapix
    for k in range(1, M_OPEN + 1):
        col = WK + k
        hcell = f"{get_column_letter(col)}${stk_hdr}"
        cfg.cell(row=stk_hdr, column=col,
                 value=(f"=IFERROR(INDEX($M$2:$M${1+S},"
                        f"MATCH({cp_lit}&\"|O\"&{k},$R$2:$R${1+S},0)),\"\")"))
        for i, mlabel in enumerate(months, start=1):
            rr = stk_hdr + i
            repwk = month_week[mlabel]
            cell = cfg.cell(row=rr, column=col,
                            value=(f"=IF({hcell}=\"\",0,SUMIFS(DashData[weighted_amount_sum],"
                                   f"DashData[pipeline_label],{cp_lit},"
                                   f"DashData[product],{prref},"
                                   f"DashData[snapshot_week_start],\"{repwk}\","
                                   f"DashData[dealstage_label],{hcell}))"))
            cell.number_format = "#,##0"

    # ===== Pipeline Dashboard sheet =====
    dash.sheet_view.showGridLines = False
    dash["A1"] = "Weighted Pipeline Dashboard"
    dash["A1"].font = TITLE_FONT
    dash["A2"] = "Pick a product, pipeline and month — KPIs and chart update automatically."
    dash["A2"].font = Font(italic=True, size=9, color="808080")

    controls = [(4, "Product", prod_list[0]),
                (5, "Pipeline", pipelines[0]),
                (6, "Month", latest_month)]
    for row_i, label, default in controls:
        lc = dash.cell(row=row_i, column=2, value=label)
        lc.font = HEADER_FONT
        vc = dash.cell(row=row_i, column=3, value=default)
        vc.font = Font(bold=True)
        vc.fill = PatternFill("solid", fgColor="FFF2CC")

    dv_specs = [(f"_cfg!$T$2:$T${1+len(prod_list)}", "C4"),
                (f"_cfg!$U$2:$U${1+NP}", "C5"),
                (f"_cfg!$S$2:$S${1+nM}", "C6")]
    for src_ref, cell in dv_specs:
        dv = DataValidation(type="list", formula1=src_ref, allow_blank=False)
        dash.add_data_validation(dv)
        dv.add(dash[cell])

    kpis = [
        (9,  "Weighted pipeline (open)",
         f"=SUMPRODUCT(_cfg!$AC${res_first}:$AC${res_last},_cfg!$AE${res_first}:$AE${res_last})", "#,##0"),
        (10, "Raw pipeline (open)",
         f"=SUMPRODUCT(_cfg!$AB${res_first}:$AB${res_last},_cfg!$AE${res_first}:$AE${res_last})", "#,##0"),
        (11, "Open deals",
         f"=SUMPRODUCT(_cfg!$AD${res_first}:$AD${res_last},_cfg!$AE${res_first}:$AE${res_last})", "#,##0"),
        (12, "Blended probability", "=IFERROR(C9/C10,0)", "0.0%"),
        (13, "Closed won (selected month)",
         f"=SUMPRODUCT(_cfg!$AB${res_first}:$AB${res_last},_cfg!$AF${res_first}:$AF${res_last})", "#,##0"),
    ]
    for row_i, label, formula, fmt in kpis:
        dash.cell(row=row_i, column=2, value=label).font = HEADER_FONT
        vc = dash.cell(row=row_i, column=3, value=formula)
        vc.number_format = fmt
        vc.font = Font(bold=True, size=12)

    dash.column_dimensions["A"].width = 3
    dash.column_dimensions["B"].width = 28
    dash.column_dimensions["C"].width = 22

    # caption above the chart — the chart is fixed to the Tapix pipeline
    cap = dash.cell(row=15, column=2,
                    value=f"{chart_pipeline} pipeline — weighted amount by stage, "
                          f"by month  (Product + Month selectors apply)")
    cap.font = Font(italic=True, size=9, color="808080")

    # chart: stacked columns — styled to match the user's Chart1_a.crtx template
    # (no chart title, value axis with fixed max + grey gridlines, data labels on,
    #  axis titles, legend at bottom, clean grey type)
    GREY_TEXT = "595959"      # Office "Text 1, lighter 35%"
    AXIS_LINE = "D9D9D9"      # Office "Text 1, lighter 85%"
    SERIES_COLORS = ["4472C4", "ED7D31", "A5A5A5", "FFC000",
                     "5B9BD5", "70AD47", "264478", "9E480E"]

    def _grey_txpr(sz=900):
        cp = CharacterProperties(sz=sz, b=False, solidFill=GREY_TEXT)
        return RichText(bodyPr=RichTextProperties(),
                        p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])

    def _grey_line():
        return GraphicalProperties(ln=LineProperties(solidFill=AXIS_LINE, w=9525))

    stk = BarChart()
    stk.type = "col"
    stk.grouping = "stacked"
    stk.overlap = 100
    stk.gapWidth = 150
    stk.title = None                       # template: autoTitleDeleted
    stk.height = 12
    stk.width = 22
    stk.add_data(Reference(cfg, min_col=WK + 1, max_col=WK + M_OPEN,
                           min_row=stk_hdr, max_row=stk_last), titles_from_data=True)
    stk.set_categories(Reference(cfg, min_col=WK, min_row=stk_first, max_row=stk_last))

    # data labels: show the value only; "#,##0;;" blanks zero-value labels
    # (empty stages like Qualify would otherwise clutter the baseline with "0")
    stk.dataLabels = DataLabelList(showVal=True, showCatName=False, showSerName=False,
                                   showLegendKey=False, showPercent=False,
                                   showBubbleSize=False, numFmt="#,##0;;")
    stk.dataLabels.txPr = _grey_txpr(800)

    # category (x) axis: visible at bottom, axis title, light-grey line, grey 9pt
    stk.x_axis.delete = False
    stk.x_axis.title = "Month"
    stk.x_axis.majorTickMark = "out"
    stk.x_axis.minorTickMark = "none"
    stk.x_axis.spPr = _grey_line()
    stk.x_axis.txPr = _grey_txpr()

    # value (y) axis: visible, light-grey gridlines.
    # NOTE: template Chart1_a pins the max at 5,000,000; against weighted data
    # (~1-2M) that leaves the chart ~60% empty, so the axis auto-fits instead.
    stk.y_axis.delete = False
    stk.y_axis.title = "Weighted amount"
    stk.y_axis.numFmt = "#,##0"
    stk.y_axis.majorTickMark = "none"
    stk.y_axis.majorGridlines = ChartLines(spPr=_grey_line())
    stk.y_axis.txPr = _grey_txpr()

    # legend at bottom, grey 9pt
    stk.legend.position = "b"
    stk.legend.overlay = False
    stk.legend.txPr = _grey_txpr()

    # series fills: Office accent palette, no bar borders
    for i, ser in enumerate(stk.series):
        gp = GraphicalProperties(solidFill=SERIES_COLORS[i % len(SERIES_COLORS)],
                                 ln=LineProperties(noFill=True))
        ser.graphicalProperties = gp

    dash.add_chart(stk, "B16")

    idx = wb.sheetnames.index("Pipeline Dashboard")
    wb.move_sheet("Pipeline Dashboard", offset=1 - idx)


def main():
    load_dotenv(ENV_PATH)

    token = os.getenv("HUBSPOT_PRIVATE_APP_TOKEN")
    if not token:
        raise RuntimeError("Chybí HUBSPOT_PRIVATE_APP_TOKEN v .env")

    OUTPUT_DIR.mkdir(exist_ok=True)
    out_xlsx = str(OUTPUT_DIR / "hubspot_deals_by_product.xlsx")

    product_label = os.getenv("PRODUCT_PROPERTY_LABEL", "Product")
    explicit_prop = os.getenv("PRODUCT_PROPERTY_NAME")  # optional

    products_interest = DEFAULT_PRODUCTS[:]

    # ✅ snapshot pro týden = pondělí (týdny jsou správně, neměníme)
    snapshot_week = week_start_iso()

    product_property_name = find_product_property_name(token, product_label, explicit_prop)
    opt_map = get_product_options_map(token, product_property_name)
    pipe_lbl, stage_lbl = get_pipelines_map(token)
    owners_map = get_owners_map(token)

    properties = [
        "dealname", "amount", "closedate", "createdate", "hs_lastmodifieddate",
        "pipeline", "dealstage",
        "hubspot_owner_id",
        product_property_name
    ]

    deals = list_all_deals(token, properties=properties)

    deal_ids = [str(d.get("id")) for d in deals if d.get("id")]
    deal_to_company = batch_read_deal_company_primary(token, deal_ids)

    company_ids = sorted({cid for cid in deal_to_company.values() if cid})
    company_names = batch_read_company_names(token, company_ids)

    rows_by_product = build_rows(
        deals=deals,
        snapshot_week_start=snapshot_week,
        product_property_name=product_property_name,
        opt_map=opt_map,
        pipeline_label=pipe_lbl,
        stage_label=stage_lbl,
        products_interest=products_interest,
        deal_to_company_id=deal_to_company,
        company_id_to_name=company_names,
        owners_map=owners_map,
    )

    # ✅ vytvoř nový soubor pokud neexistuje
    if not os.path.exists(out_xlsx):
        wb = create_new_workbook(out_xlsx, products_interest)
    else:
        wb = load_workbook(out_xlsx)
        if "Summary" not in wb.sheetnames:
            wb.create_sheet("Summary", 0)
        for p in products_interest:
            sh = excel_safe_sheet_name(p)
            if sh not in wb.sheetnames:
                wb.create_sheet(sh)
            ensure_sheet_headers(wb[sh])

    # ✅ NOVÉ: týden se jednou zapíše, pak už se nepřepisuje
    for product, rows in rows_by_product.items():
        sh_name = excel_safe_sheet_name(product)
        ws = wb[sh_name]
        ensure_sheet_headers(ws)
        replace_rows_for_snapshot(ws, rows, snapshot_week)

    rewrite_summary_sheet(wb, products_interest)
    build_dashboard(wb, products_interest)

    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    wb.save(out_xlsx)
    print(f"✅ Hotovo: {out_xlsx}")
    print(f" snapshot_week_start={snapshot_week}, deals_fetched={len(deals)}, companies_found={len(company_ids)}")


if __name__ == "__main__":
    main()
